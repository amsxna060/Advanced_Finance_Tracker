"""
Excel Export Service using openpyxl
Generates Excel workbooks with multiple sheets and formatting
"""
from datetime import datetime
from typing import List, Dict, Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generate various financial reports in Excel format"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="1e40af")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _format_currency(self, amount: float) -> float:
        """Return amount as float for Excel formatting"""
        return float(amount) if amount else 0.0
    
    def _format_date(self, date) -> str:
        """Format date"""
        if isinstance(date, str):
            return date
        return date.strftime('%Y-%m-%d') if date else ''
    
    def _apply_header_style(self, ws, row_idx: int, max_col: int):
        """Apply header styling to a row"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _apply_border(self, ws, start_row: int, end_row: int, max_col: int):
        """Apply borders to a range"""
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = self.border
    
    def _auto_size_columns(self, ws, max_col: int):
        """Auto-size columns based on content"""
        for col in range(1, max_col + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def generate_comprehensive_report(
        self,
        summary_data: Dict[str, Any],
        loans: List[Dict[str, Any]],
        properties: List[Dict[str, Any]],
        partnerships: List[Dict[str, Any]],
        expenses: List[Dict[str, Any]]
    ) -> BytesIO:
        """Generate comprehensive Excel report with multiple sheets"""
        buffer = BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Portfolio Summary
        ws_summary = wb.create_sheet("Portfolio Summary")
        self._create_summary_sheet(ws_summary, summary_data)
        
        # Sheet 2: Loans
        ws_loans = wb.create_sheet("Loans")
        self._create_loans_sheet(ws_loans, loans)
        
        # Sheet 3: Properties
        ws_properties = wb.create_sheet("Properties")
        self._create_properties_sheet(ws_properties, properties)
        
        # Sheet 4: Partnerships
        ws_partnerships = wb.create_sheet("Partnerships")
        self._create_partnerships_sheet(ws_partnerships, partnerships)
        
        # Sheet 5: Expenses
        ws_expenses = wb.create_sheet("Expenses")
        self._create_expenses_sheet(ws_expenses, expenses)
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, ws, data: Dict[str, Any]):
        """Create portfolio summary sheet"""
        # Title
        ws['A1'] = "PORTFOLIO SUMMARY REPORT"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:B2')
        
        # Lending Portfolio
        row = 4
        ws[f'A{row}'] = "LENDING PORTFOLIO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        self._apply_header_style(ws, row, 2)
        
        lending_metrics = [
            ("Total Lent Out", self._format_currency(data.get('total_lent_out', 0))),
            ("Outstanding Receivables", self._format_currency(data.get('total_outstanding_receivable', 0))),
            ("Expected This Month", self._format_currency(data.get('expected_this_month', 0))),
            ("Total Overdue", self._format_currency(data.get('total_overdue', 0))),
        ]
        
        for metric, value in lending_metrics:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '₹#,##0.00'
        
        # Borrowing Portfolio
        row += 2
        ws[f'A{row}'] = "BORROWING PORTFOLIO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        self._apply_header_style(ws, row, 2)
        
        borrowing_metrics = [
            ("Total Borrowed", self._format_currency(data.get('total_borrowed', 0))),
            ("Outstanding Payables", self._format_currency(data.get('total_outstanding_payable', 0))),
        ]
        
        for metric, value in borrowing_metrics:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '₹#,##0.00'
        
        # Net Position
        row += 2
        ws[f'A{row}'] = "NET POSITION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Net Position"
        ws[f'B{row}'] = self._format_currency(data.get('net_position', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        # Other Investments
        row += 2
        ws[f'A{row}'] = "OTHER INVESTMENTS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Active Property Deals"
        ws[f'B{row}'] = data.get('active_property_deals', 0)
        
        row += 1
        ws[f'A{row}'] = "Active Partnerships"
        ws[f'B{row}'] = data.get('active_partnerships', 0)
        
        self._auto_size_columns(ws, 2)
    
    def _create_loans_sheet(self, ws, loans: List[Dict[str, Any]]):
        """Create loans sheet"""
        headers = ['ID', 'Contact', 'Type', 'Principal', 'Rate %', 'Outstanding', 
                   'Status', 'Start Date', 'Tenure (Months)']
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        self._apply_header_style(ws, 1, len(headers))
        
        for row, loan in enumerate(loans, start=2):
            ws.cell(row=row, column=1, value=loan.get('id'))
            ws.cell(row=row, column=2, value=loan.get('contact_name', 'N/A'))
            ws.cell(row=row, column=3, value=loan.get('loan_type', 'N/A'))
            ws.cell(row=row, column=4, value=self._format_currency(loan.get('principal', 0)))
            ws.cell(row=row, column=5, value=loan.get('interest_rate', 0))
            ws.cell(row=row, column=6, value=self._format_currency(loan.get('outstanding', 0)))
            ws.cell(row=row, column=7, value=loan.get('status', 'N/A'))
            ws.cell(row=row, column=8, value=self._format_date(loan.get('start_date')))
            ws.cell(row=row, column=9, value=loan.get('tenure_months', 0))
            
            # Format currency columns
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
        
        self._apply_border(ws, 1, len(loans) + 1, len(headers))
        self._auto_size_columns(ws, len(headers))
    
    def _create_properties_sheet(self, ws, properties: List[Dict[str, Any]]):
        """Create properties sheet"""
        headers = ['ID', 'Title', 'Location', 'Total Investment', 'Current Value', 
                   'Status', 'Purchase Date', 'Exit Date']
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        self._apply_header_style(ws, 1, len(headers))
        
        for row, prop in enumerate(properties, start=2):
            ws.cell(row=row, column=1, value=prop.get('id'))
            ws.cell(row=row, column=2, value=prop.get('title', 'N/A'))
            ws.cell(row=row, column=3, value=prop.get('location', 'N/A'))
            ws.cell(row=row, column=4, value=self._format_currency(prop.get('total_investment', 0)))
            ws.cell(row=row, column=5, value=self._format_currency(prop.get('current_value', 0)))
            ws.cell(row=row, column=6, value=prop.get('status', 'N/A'))
            ws.cell(row=row, column=7, value=self._format_date(prop.get('purchase_date')))
            ws.cell(row=row, column=8, value=self._format_date(prop.get('exit_date')))
            
            # Format currency columns
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        
        self._apply_border(ws, 1, len(properties) + 1, len(headers))
        self._auto_size_columns(ws, len(headers))
    
    def _create_partnerships_sheet(self, ws, partnerships: List[Dict[str, Any]]):
        """Create partnerships sheet"""
        headers = ['ID', 'Name', 'Type', 'Total Capital', 'Your Share %', 
                   'Your Investment', 'Status', 'Start Date']
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        self._apply_header_style(ws, 1, len(headers))
        
        for row, partnership in enumerate(partnerships, start=2):
            ws.cell(row=row, column=1, value=partnership.get('id'))
            ws.cell(row=row, column=2, value=partnership.get('name', 'N/A'))
            ws.cell(row=row, column=3, value=partnership.get('partnership_type', 'N/A'))
            ws.cell(row=row, column=4, value=self._format_currency(partnership.get('total_capital', 0)))
            ws.cell(row=row, column=5, value=partnership.get('your_share_percentage', 0))
            ws.cell(row=row, column=6, value=self._format_currency(partnership.get('your_investment', 0)))
            ws.cell(row=row, column=7, value=partnership.get('status', 'N/A'))
            ws.cell(row=row, column=8, value=self._format_date(partnership.get('start_date')))
            
            # Format currency columns
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
        
        self._apply_border(ws, 1, len(partnerships) + 1, len(headers))
        self._auto_size_columns(ws, len(headers))
    
    def _create_expenses_sheet(self, ws, expenses: List[Dict[str, Any]]):
        """Create expenses sheet"""
        headers = ['ID', 'Date', 'Category', 'Amount', 'Linked Type', 
                   'Payment Mode', 'Description']
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        self._apply_header_style(ws, 1, len(headers))
        
        total_amount = 0
        for row, expense in enumerate(expenses, start=2):
            ws.cell(row=row, column=1, value=expense.get('id'))
            ws.cell(row=row, column=2, value=self._format_date(expense.get('expense_date')))
            ws.cell(row=row, column=3, value=expense.get('category', 'N/A'))
            amount = self._format_currency(expense.get('amount', 0))
            ws.cell(row=row, column=4, value=amount)
            ws.cell(row=row, column=5, value=expense.get('linked_type', 'N/A'))
            ws.cell(row=row, column=6, value=expense.get('payment_mode', 'N/A'))
            ws.cell(row=row, column=7, value=expense.get('description', ''))
            
            # Format currency column
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            total_amount += amount
        
        # Add total row
        if expenses:
            total_row = len(expenses) + 2
            ws.cell(row=total_row, column=3, value="TOTAL")
            ws.cell(row=total_row, column=4, value=total_amount)
            ws.cell(row=total_row, column=3).font = Font(bold=True)
            ws.cell(row=total_row, column=4).font = Font(bold=True)
            ws.cell(row=total_row, column=4).number_format = '₹#,##0.00'
        
        self._apply_border(ws, 1, len(expenses) + 1, len(headers))
        self._auto_size_columns(ws, len(headers))
