"""
Reports Router - PDF and Excel report generation endpoints
"""
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user, require_module
from app.models.user import User
from app.models.loan import Loan, LoanPayment
from app.models.property_deal import PropertyDeal, PropertyTransaction
from app.models.partnership import Partnership, PartnershipTransaction
from app.models.expense import Expense
from app.services.interest import calculate_outstanding
from app.services.pdf_generator import PDFReportGenerator
from app.services.excel_generator import ExcelReportGenerator
# F5: share the dashboard's txn-type vocabulary (new + legacy) so reports and
# dashboard can never disagree on what counts as property/partnership money.
from app.routers.dashboard import (
    INFLOW_PROPERTY_TXN_TYPES,
    OUTFLOW_PROPERTY_TXN_TYPES,
    INFLOW_PARTNERSHIP_TXN_TYPES,
)

router = APIRouter(prefix="/api/reports", tags=["reports"], dependencies=[Depends(require_module("reports"))])

# H-REP-2: do NOT create module-level singletons — they hold mutable state
# (open file handles, in-progress workbooks) that would be shared across
# concurrent requests. Instantiate fresh instances per request instead.
# pdf_generator and excel_generator are now created inside each endpoint.


def _d(value) -> Decimal:
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


@router.get("/loan-statement/{loan_id}")
def generate_loan_statement(
    loan_id: int,
    format: str = Query("pdf", regex="^(pdf|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # C-REP-1: filter out soft-deleted loans
    loan = db.query(Loan).filter(Loan.id == loan_id, Loan.is_deleted == False).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    # C-REP-1: exclude voided payments from the statement
    payments = (
        db.query(LoanPayment)
        .filter(LoanPayment.loan_id == loan_id, LoanPayment.is_voided == False)
        .order_by(LoanPayment.payment_date.asc())
        .all()
    )

    outstanding_data = calculate_outstanding(loan.id, date.today(), db)

    loan_data = {
        "id": loan.id,
        "contact_name": loan.contact.name if loan.contact else "N/A",
        "principal": float(_d(loan.principal_amount)),
        "interest_rate": float(_d(loan.interest_rate)),
        "start_date": loan.disbursed_date,
        "tenure_months": loan.tenure_months or 0,
        "status": loan.status,
        "outstanding": float(_d(outstanding_data["total_outstanding"])),
        "principal_outstanding": float(_d(outstanding_data["principal_outstanding"])),
        "interest_outstanding": float(_d(outstanding_data["interest_outstanding"])),
        "total_paid": float(sum(_d(p.amount_paid) for p in payments)),
        "overdue_amount": float(_d(outstanding_data["interest_outstanding"])),
    }

    payment_data = [
        {
            "payment_date": p.payment_date,
            "payment_type": p.payment_mode or "payment",
            "principal_amount": float(_d(p.allocated_to_principal)),
            "interest_amount": float(_d(p.allocated_to_current_interest) + _d(p.allocated_to_overdue_interest)),
            "total_amount": float(_d(p.amount_paid)),
            # H-REP-3: compute running outstanding balance instead of always returning 0
            "outstanding_after": 0,  # filled below
        }
        for p in payments
    ]
    # H-REP-3: fill running outstanding per payment using calculate_outstanding per date
    for i, p in enumerate(payments):
        as_of = calculate_outstanding(loan.id, p.payment_date, db)
        payment_data[i]["outstanding_after"] = float(_d(as_of["total_outstanding"]))

    if format == "pdf":
        buffer = PDFReportGenerator().generate_loan_statement(loan_data, payment_data)
        filename = f"loan_statement_{loan_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        media_type = "application/pdf"
    else:
        buffer = _loan_excel(loan_data, payment_data)
        filename = f"loan_statement_{loan_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        # M-REP-5: quote filename in Content-Disposition
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _loan_excel(loan_data: dict, payment_data: list) -> BytesIO:
    from openpyxl import Workbook
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Statement"
    ws["A1"] = "Loan Statement"
    ws["A2"] = f"Loan ID: {loan_data['id']}"
    ws["A3"] = f"Contact: {loan_data['contact_name']}"
    ws["A4"] = f"Principal: {loan_data['principal']}"
    ws["A5"] = f"Outstanding: {loan_data['outstanding']}"
    ws["A6"] = f"Total Paid: {loan_data['total_paid']}"
    ws["A7"] = f"Status: {loan_data['status']}"
    ws["A9"] = "Payment History"
    headers = ["Date", "Mode", "Principal", "Interest", "Total Paid", "Balance After"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=10, column=col, value=h)
    for row, p in enumerate(payment_data, start=11):
        ws.cell(row=row, column=1, value=str(p["payment_date"]))
        ws.cell(row=row, column=2, value=p["payment_type"])
        ws.cell(row=row, column=3, value=p["principal_amount"])
        ws.cell(row=row, column=4, value=p["interest_amount"])
        ws.cell(row=row, column=5, value=p["total_amount"])
        ws.cell(row=row, column=6, value=p["outstanding_after"])
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/portfolio-summary")
def generate_portfolio_summary(
    format: str = Query("pdf", regex="^(pdf|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # C-REP-1: exclude soft-deleted loans from portfolio summary
    # F6: no created_by scoping — the rest of the app (dashboard D1 fix, lists,
    # analytics) shows all data, so a per-creator report contradicted the screens.
    active_loans = db.query(Loan).filter(Loan.status == "active", Loan.is_deleted == False).all()

    total_lent_out = Decimal("0")
    total_outstanding_receivable = Decimal("0")
    total_borrowed = Decimal("0")
    total_outstanding_payable = Decimal("0")
    # H-REP-4: compute real overdue and expected-this-month figures
    total_overdue = Decimal("0")
    today = date.today()
    this_month_start = today.replace(day=1)
    # next month start to bound expected-this-month
    if today.month == 12:
        next_month_start = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month_start = today.replace(month=today.month + 1, day=1)
    expected_this_month = Decimal("0")

    for loan in active_loans:
        principal = _d(loan.principal_amount)
        outstanding = calculate_outstanding(loan.id, today, db)
        total_due = _d(outstanding["total_outstanding"])
        overdue = _d(outstanding.get("overdue_interest", 0))
        if loan.loan_direction == "given":
            total_lent_out += principal
            total_outstanding_receivable += total_due
            total_overdue += overdue
            # expected this month: next EMI amount if due this month
            if loan.emi_amount and loan.emi_day_of_month:
                try:
                    emi_due = today.replace(day=int(loan.emi_day_of_month))
                    if this_month_start <= emi_due < next_month_start:
                        expected_this_month += _d(loan.emi_amount)
                except ValueError:
                    pass
        else:
            total_borrowed += principal
            total_outstanding_payable += total_due

    active_property_deals = db.query(PropertyDeal).filter(PropertyDeal.is_deleted == False).count()
    active_partnerships = db.query(Partnership).filter(Partnership.status == "active", Partnership.is_deleted == False).count()

    summary_data = {
        "total_lent_out": float(total_lent_out),
        "total_outstanding_receivable": float(total_outstanding_receivable),
        "total_borrowed": float(total_borrowed),
        "total_outstanding_payable": float(total_outstanding_payable),
        "net_position": float(total_outstanding_receivable - total_outstanding_payable),
        # H-REP-4: actual computed values instead of hardcoded 0
        "expected_this_month": float(expected_this_month),
        "total_overdue": float(total_overdue),
        "active_property_deals": active_property_deals,
        "active_partnerships": active_partnerships,
    }

    if format == "pdf":
        buffer = PDFReportGenerator().generate_portfolio_summary(summary_data)
        filename = f"portfolio_summary_{datetime.now().strftime('%Y%m%d')}.pdf"
        media_type = "application/pdf"
    else:
        loans_list = []
        for loan in active_loans:
            outstanding = calculate_outstanding(loan.id, date.today(), db)
            loans_list.append({
                "id": loan.id,
                "contact_name": loan.contact.name if loan.contact else "N/A",
                "loan_type": loan.loan_direction,
                "principal": float(_d(loan.principal_amount)),
                "interest_rate": float(_d(loan.interest_rate)),
                "outstanding": float(_d(outstanding["total_outstanding"])),
                "status": loan.status,
                "start_date": loan.disbursed_date,
                "tenure_months": loan.tenure_months or 0,
            })

        properties_list = [
            {
                "id": p.id,
                "title": p.title,
                "location": p.location or "",
                "total_investment": float(_d(p.purchase_price or p.advance_paid)),
                "current_value": float(_d(p.sale_price or p.total_buyer_value or p.total_seller_value)),
                "status": p.status,
                "purchase_date": p.deal_locked_date,
                "exit_date": p.actual_registry_date,
            }
            # F7: was unfiltered — mixed deleted properties into the Excel while
            # every other section of the same report was filtered
            for p in db.query(PropertyDeal).filter(PropertyDeal.is_deleted == False).all()
        ]

        partnerships_list = [
            {
                "id": pr.id,
                "name": pr.title,
                "partnership_type": "partnership",
                "total_capital": float(_d(pr.total_deal_value)),
                "your_share_percentage": float(_d(pr.our_share_percentage)),
                "your_investment": float(_d(pr.our_investment)),
                "status": pr.status,
                "start_date": pr.start_date,
            }
            for pr in db.query(Partnership).filter(Partnership.is_deleted == False).all()
        ]

        expenses_list = [
            {
                "id": e.id,
                "expense_date": e.expense_date,
                "category": e.category,
                "amount": float(_d(e.amount)),
                "linked_type": e.linked_type or "",
                "payment_mode": e.payment_mode or "",
                "description": e.description or "",
            }
            for e in db.query(Expense).filter(Expense.is_deleted == False).all()
        ]

        buffer = ExcelReportGenerator().generate_comprehensive_report(
            summary_data, loans_list, properties_list, partnerships_list, expenses_list
        )
        filename = f"portfolio_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/profit-loss")
def generate_pnl_report(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    format: str = Query("pdf", regex="^(pdf|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # M-VAL-7: catch bad date strings and return 422 instead of crashing with 500
    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Invalid end_date: '{end_date}'. Use YYYY-MM-DD format.")
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else end - timedelta(days=30)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Invalid start_date: '{start_date}'. Use YYYY-MM-DD format.")

    # Income
    income_data = []
    received_payments = (
        db.query(LoanPayment)
        .join(Loan)
        .filter(
            Loan.loan_direction == "given",
            Loan.is_deleted == False,
            # C-REP-1: exclude voided payments from profit-loss income
            LoanPayment.is_voided == False,
            LoanPayment.payment_date.between(start, end),
        )
        .all()
    )
    total_interest_received = sum(
        _d(p.allocated_to_current_interest) + _d(p.allocated_to_overdue_interest)
        for p in received_payments
    )
    if total_interest_received > 0:
        income_data.append({"category": "Interest Income", "amount": float(total_interest_received)})

    # F5: use the shared (new + legacy) type vocabulary and exclude voided rows
    prop_inflow = sum(
        _d(t.amount)
        for t in db.query(PropertyTransaction).filter(
            PropertyTransaction.txn_date.between(start, end),
            PropertyTransaction.txn_type.in_(INFLOW_PROPERTY_TXN_TYPES),
            PropertyTransaction.is_voided == False,
        ).all()
    )
    if prop_inflow > 0:
        income_data.append({"category": "Property Income", "amount": float(prop_inflow)})

    partnership_income = sum(
        _d(t.amount)
        for t in db.query(PartnershipTransaction).filter(
            PartnershipTransaction.txn_date.between(start, end),
            PartnershipTransaction.txn_type.in_(INFLOW_PARTNERSHIP_TXN_TYPES),
            PartnershipTransaction.is_voided == False,
        ).all()
    )
    if partnership_income > 0:
        income_data.append({"category": "Partnership Income", "amount": float(partnership_income)})

    # Expenses
    expense_data = []
    # C-REP-1: exclude soft-deleted expenses from profit-loss
    expenses = db.query(Expense).filter(
        Expense.expense_date.between(start, end),
        Expense.is_deleted == False,
    ).all() if hasattr(Expense, 'is_deleted') else db.query(Expense).filter(
        Expense.expense_date.between(start, end),
    ).all()
    by_category: dict = {}
    for e in expenses:
        cat = (e.category or "other").capitalize()
        by_category[cat] = by_category.get(cat, Decimal("0")) + _d(e.amount)
    for cat, amt in by_category.items():
        expense_data.append({"category": cat, "amount": float(amt)})

    paid_payments = (
        db.query(LoanPayment)
        .join(Loan)
        .filter(
            Loan.loan_direction == "taken",
            Loan.is_deleted == False,
            # C-REP-1: exclude voided payments from expenses
            LoanPayment.is_voided == False,
            LoanPayment.payment_date.between(start, end),
        )
        .all()
    )
    total_interest_paid = sum(
        _d(p.allocated_to_current_interest) + _d(p.allocated_to_overdue_interest)
        for p in paid_payments
    )
    if total_interest_paid > 0:
        expense_data.append({"category": "Interest Expense", "amount": float(total_interest_paid)})

    prop_outflow = sum(
        _d(t.amount)
        for t in db.query(PropertyTransaction).filter(
            PropertyTransaction.txn_date.between(start, end),
            PropertyTransaction.txn_type.in_(OUTFLOW_PROPERTY_TXN_TYPES),
            PropertyTransaction.is_voided == False,
        ).all()
    )
    if prop_outflow > 0:
        expense_data.append({"category": "Property Payments", "amount": float(prop_outflow)})

    if format == "pdf":
        buffer = PDFReportGenerator().generate_pnl_report(
            datetime.combine(start, datetime.min.time()),
            datetime.combine(end, datetime.min.time()),
            income_data,
            expense_data,
        )
        filename = f"pnl_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
        media_type = "application/pdf"
    else:
        buffer = _pnl_excel(start, end, income_data, expense_data)
        filename = f"pnl_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        # M-REP-5: quote filename so Content-Disposition header is valid RFC 6266
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pnl_excel(start: date, end: date, income_data: list, expense_data: list) -> BytesIO:
    from openpyxl import Workbook
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Statement"
    ws["A1"] = "PROFIT & LOSS STATEMENT"
    ws["A2"] = f"Period: {start} to {end}"
    row = 4
    ws[f"A{row}"] = "INCOME"
    for item in income_data:
        row += 1
        ws[f"A{row}"] = item["category"]
        ws[f"B{row}"] = item["amount"]
    total_income = sum(i["amount"] for i in income_data)
    row += 1
    ws[f"A{row}"] = "Total Income"
    ws[f"B{row}"] = total_income
    row += 2
    ws[f"A{row}"] = "EXPENSES"
    for item in expense_data:
        row += 1
        ws[f"A{row}"] = item["category"]
        ws[f"B{row}"] = item["amount"]
    total_expense = sum(e["amount"] for e in expense_data)
    row += 1
    ws[f"A{row}"] = "Total Expenses"
    ws[f"B{row}"] = total_expense
    row += 2
    ws[f"A{row}"] = "Net Profit / Loss"
    ws[f"B{row}"] = total_income - total_expense
    wb.save(buffer)
    buffer.seek(0)
    return buffer
